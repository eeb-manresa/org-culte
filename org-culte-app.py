import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime

st.set_page_config(page_title="Ordre Culte EEBM", page_icon="⛪", layout="wide")

EXCEL_FILE = "planificacion.xlsx"

def load_wb_data():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    # 1. Cargar Planificación
    ws_plan = wb["Planificación"]
    plan_data = []
    for row in range(5, ws_plan.max_row + 1):
        row_vals = [ws_plan.cell(row=row, column=col).value for col in range(1, 17)]
        if row_vals[0] is not None:
            plan_data.append(row_vals)
            
    headers_plan = [
        "Fecha", "Culto / Nota", "Puerta 1", "Puerta 2", "Entrada 1", "Entrada 2",
        "Presidencia", "Predicación", "Alabanza", "Sonido", "Proyección",
        "Santa Cena 1", "Santa Cena 2", "Santa Cena 3", "Ofrenda 1", "Ofrenda 2"
    ]
    df_plan = pd.DataFrame(plan_data, columns=headers_plan)
    
    if not df_plan.empty:
        df_plan["Fecha"] = pd.to_datetime(df_plan["Fecha"], errors="coerce")
    
    # 2. Cargar Voluntarios
    ws_vol = wb["Voluntarios"]
    vol_dict = {}
    for col in range(1, ws_vol.max_column + 1):
        area_name = ws_vol.cell(row=1, column=col).value
        if area_name:
            names = []
            for r in range(2, ws_vol.max_row + 1):
                val = ws_vol.cell(row=r, column=col).value
                if val:
                    names.append(val)
            vol_dict[area_name] = names
            
    wb.close()
    return df_plan, vol_dict

df_plan, vol_dict = load_wb_data()

st.title("⛪ Ordre Culte EEBM")
st.markdown("Aplicació web interactiva per a la coordinació del servei dominical.")

tab1, tab2, tab3, tab4 = st.tabs([
    "📊 Resum per Culte", 
    "✏️ Assignar / Editar Culte", 
    "📅 Planificació General", 
    "👥 Gestionar Directori"
])

# --- PESTAÑA 1: RESUMEN EJECUTIVO ---
with tab1:
    st.subheader("Assignació Diumenge")
    if not df_plan.empty:
        dates_list = sorted(df_plan["Fecha"].dropna().dt.date.unique())
        selected_date = st.selectbox("Selecciona la data del culte:", dates_list, key="resumen_date")
        
        if selected_date:
            row_data = df_plan[df_plan["Fecha"].dt.date == selected_date].iloc[0]
            st.markdown(f"### Culte: {row_data['Culto / Nota']} ({selected_date.strftime('%d/%m/%Y')})")
            
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("#### 🚪 Benvinguda")
                st.info(f"""**Porta 1:** {row_data['Puerta 1'] or '-'}
**Porta 2:** {row_data['Puerta 2'] or '-'}
**Entrada 1:** {row_data['Entrada 1'] or '-'}
**Entrada 2:** {row_data['Entrada 2'] or '-'}""")
                
                st.markdown("#### 📖 Direcció i Paraula")
                st.success(f"""**Presidència:** {row_data['Presidencia'] or '-'}
**Predicació:** {row_data['Predicación'] or '-'}""")
                
            with col2:
                st.markdown("#### 🎵 Tècnica i Música")
                st.warning(f"""**Alabança:** {row_data['Alabanza'] or '-'}
**So:** {row_data['Sonido'] or '-'}
**Projecció:** {row_data['Proyección'] or '-'}""")
                
                st.markdown("#### 🍞 Ordenances i Ofrena")
                st.error(f"""**Santa Cena:** {row_data['Santa Cena 1'] or '-'}, {row_data['Santa Cena 2'] or '-'}, {row_data['Santa Cena 3'] or '-'}
**Ofrena:** {row_data['Ofrenda 1'] or '-'}, {row_data['Ofrenda 2'] or '-'}""")
    else:
        st.warning("No hi ha dades a la planificació.")

# --- PESTAÑA 2: ASIGNAR / EDITAR DESDE LA WEB ---
with tab2:
    st.subheader("Assignació de Voluntaris per Culte")
    
    with st.expander("➕ Crear un nou culte en una altra data"):
        with st.form("form_create_culte"):
            new_culte_date = st.date_input("Selecciona la data del nou culte")
            new_culte_nota = st.text_input("Tipus de Culte / Nota", value="Culte General")
            btn_create = st.form_submit_button("Crear Culte")
            
            if btn_create:
                wb_c = openpyxl.load_workbook(EXCEL_FILE)
                ws_c = wb_c["Planificación"]
                next_r = 5
                while ws_c.cell(row=next_r, column=1).value is not None:
                    next_r += 1
                
                ws_c.cell(row=next_r, column=1, value=pd.to_datetime(new_culte_date))
                ws_c.cell(row=next_r, column=2, value=new_culte_nota)
                wb_c.save(EXCEL_FILE)
                wb_c.close()
                st.success(f"S'ha creat el culte per al dia {new_culte_date.strftime('%d/%m/%Y')}. Recarrega la pàgina.")
                st.rerun()

    if not df_plan.empty:
        dates_str = [d.strftime('%Y-%m-%d') for d in sorted(df_plan["Fecha"].dropna().dt.date.unique())]
        chosen_date_str = st.selectbox("Selecciona data a modificar:", dates_str, key="edit_date")
        
        wb_edit = openpyxl.load_workbook(EXCEL_FILE)
        ws_p = wb_edit["Planificación"]
        
        target_row = None
        for r in range(5, ws_p.max_row + 1):
            cell_val = ws_p.cell(row=r, column=1).value
            if cell_val:
                cell_date_str = pd.to_datetime(cell_val).strftime('%Y-%m-%d')
                if cell_date_str == chosen_date_str:
                    target_row = r
                    break
                
        if target_row:
            current_nota = ws_p.cell(row=target_row, column=2).value or "Culte General"
            
            with st.form(key="form_asignacion"):
                st.markdown(f"**Modificant el culte del dia: {chosen_date_str}**")
                nota_input = st.text_input("Tipus de Culte / Nota", value=current_nota)
                
                l_puerta = [""] + vol_dict.get("Puerta", [])
                l_entrada = [""] + vol_dict.get("Entrada", [])
                l_presi = [""] + vol_dict.get("Presidencia", [])
                l_pred = [""] + vol_dict.get("Predicación", [])
                l_alab = [""] + vol_dict.get("Alabanza", [])
                l_son = [""] + vol_dict.get("Sonido", [])
                l_proy = [""] + vol_dict.get("Proyección", [])
                l_cena = [""] + vol_dict.get("Santa Cena", [])
                l_ofer = [""] + vol_dict.get("Ofrena", vol_dict.get("Ofrenda", []))
                
                col_a, col_b = st.columns(2)
                
                with col_a:
                    st.markdown("#### 🚪 Benvinguda")
                    p1 = st.selectbox("Porta 1", l_puerta, index=l_puerta.index(ws_p.cell(row=target_row, column=3).value) if ws_p.cell(row=target_row, column=3).value in l_puerta else 0)
                    p2 = st.selectbox("Porta 2", l_puerta, index=l_puerta.index(ws_p.cell(row=target_row, column=4).value) if ws_p.cell(row=target_row, column=4).value in l_puerta else 0)
                    e1 = st.selectbox("Entrada 1", l_entrada, index=l_entrada.index(ws_p.cell(row=target_row, column=5).value) if ws_p.cell(row=target_row, column=5).value in l_entrada else 0)
                    e2 = st.selectbox("Entrada 2", l_entrada, index=l_entrada.index(ws_p.cell(row=target_row, column=6).value) if ws_p.cell(row=target_row, column=6).value in l_entrada else 0)
                    
                    st.markdown("#### 📖 Paraula")
                    pres = st.selectbox("Presidència", l_presi, index=l_presi.index(ws_p.cell(row=target_row, column=7).value) if ws_p.cell(row=target_row, column=7).value in l_presi else 0)
                    pred = st.selectbox("Predicació", l_pred, index=l_pred.index(ws_p.cell(row=target_row, column=8).value) if ws_p.cell(row=target_row, column=8).value in l_pred else 0)

                with col_b:
                    st.markdown("#### 🎵 Tècnica i Música")
                    alab = st.selectbox("Alabança", l_alab, index=l_alab.index(ws_p.cell(row=target_row, column=9).value) if ws_p.cell(row=target_row, column=9).value in l_alab else 0)
                    son = st.selectbox("So", l_son, index=l_son.index(ws_p.cell(row=target_row, column=10).value) if ws_p.cell(row=target_row, column=10).value in l_son else 0)
                    proy = st.selectbox("Projecció", l_proy, index=l_proy.index(ws_p.cell(row=target_row, column=11).value) if ws_p.cell(row=target_row, column=11).value in l_proy else 0)
                    
                    st.markdown("#### 🍞 Ordenances i Ofrena")
                    sc1 = st.selectbox("Santa Cena 1", l_cena, index=l_cena.index(ws_p.cell(row=target_row, column=12).value) if ws_p.cell(row=target_row, column=12).value in l_cena else 0)
                    sc2 = st.selectbox("Santa Cena 2", l_cena, index=l_cena.index(ws_p.cell(row=target_row, column=13).value) if ws_p.cell(row=target_row, column=13).value in l_cena else 0)
                    sc3 = st.selectbox("Santa Cena 3", l_cena, index=l_cena.index(ws_p.cell(row=target_row, column=14).value) if ws_p.cell(row=target_row, column=14).value in l_cena else 0)
                    of1 = st.selectbox("Ofrena 1", l_ofer, index=l_ofer.index(ws_p.cell(row=target_row, column=15).value) if ws_p.cell(row=target_row, column=15).value in l_ofer else 0)
                    of2 = st.selectbox("Ofrena 2", l_ofer, index=l_ofer.index(ws_p.cell(row=target_row, column=16).value) if ws_p.cell(row=target_row, column=16).value in l_ofer else 0)

                submitted = st.form_submit_button("💾 Guardar Assignació")
                if submitted:
                    ws_p.cell(row=target_row, column=2, value=nota_input)
                    ws_p.cell(row=target_row, column=3, value=p1)
                    ws_p.cell(row=target_row, column=4, value=p2)
                    ws_p.cell(row=target_row, column=5, value=e1)
                    ws_p.cell(row=target_row, column=6, value=e2)
                    ws_p.cell(row=target_row, column=7, value=pres)
                    ws_p.cell(row=target_row, column=8, value=pred)
                    ws_p.cell(row=target_row, column=9, value=alab)
                    ws_p.cell(row=target_row, column=10, value=son)
                    ws_p.cell(row=target_row, column=11, value=proy)
                    ws_p.cell(row=target_row, column=12, value=sc1)
                    ws_p.cell(row=target_row, column=13, value=sc2)
                    ws_p.cell(row=target_row, column=14, value=sc3)
                    ws_p.cell(row=target_row, column=15, value=of1)
                    ws_p.cell(row=target_row, column=16, value=of2)
                    
                    wb_edit.save(EXCEL_FILE)
                    st.success("¡Assignació actualitzada correctament! Recarrega la pàgina.")
        wb_edit.close()

# --- PESTAÑA 3: PLANIFICACIÓN GENERAL ---
with tab3:
    st.subheader("Taula Completa de Planificació")
    st.dataframe(df_plan, use_container_width=True)

# --- PESTAÑA 4: GESTIÓN DE DIRECTORIO DE VOLUNTARIOS ---
with tab4:
    st.subheader("Directori de Voluntaris per Àrea")
    st.markdown("Afegeix o elimina voluntaris de forma ràpida.")
    
    col_v1, col_v2 = st.columns(2)
    
    with col_v1:
        st.markdown("### ➕ Afegir Voluntari")
        with st.form("form_add_vol"):
            wb_v = openpyxl.load_workbook(EXCEL_FILE)
            ws_v = wb_v["Voluntarios"]
            
            area_cols = {}
            for col in range(1, ws_v.max_column + 1):
                header = ws_v.cell(row=1, column=col).value
                if header:
                    area_cols[header] = col
            
            selected_area = st.selectbox("Selecciona l'Àrea", list(area_cols.keys()))
            new_name = st.text_input("Nom i Cognoms del Voluntari")
            
            add_btn = st.form_submit_button("Afegir al Directori")
            if add_btn and new_name:
                col_idx = area_cols[selected_area]
                next_r = 2
                while ws_v.cell(row=next_r, column=col_idx).value is not None:
                    next_r += 1
                ws_v.cell(row=next_r, column=col_idx, value=new_name)
                wb_v.save(EXCEL_FILE)
                st.success(f"S'ha afegit a {new_name} a {selected_area}.")
                wb_v.close()
                st.rerun()
            wb_v.close()

    with col_v2:
        st.markdown("### 📋 Llistat Actual")
        for area, names in vol_dict.items():
            with st.expander(f"🔹 {area} ({len(names)} voluntaris)"):
                for name in names:
                    st.text(f"• {name}")